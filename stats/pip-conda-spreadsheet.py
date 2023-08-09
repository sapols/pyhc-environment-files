import requests
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_core_pyhc_packages():
    """
    TODO: consider scraping this from projects_core.yml online?
    :return: A list of the core PyHC package names.
    """
    return ["hapiclient", "kamodo", "plasmapy", "pysat", "pyspedas", "spacepy", "sunpy"]


def get_other_pyhc_packages():
    """
    TODO: consider scraping this from projects.yml online?
    :return: A list of the other non-core PyHC package names.
    """
    return ["acemag", "afino", "aeindex", "geodata", "gimamag", "polan", "pygemini", "pyglow", "python-magnetosphere", "sami2py", "scanning-doppler-interferometer", "aacgmv2", "aiapy", "aidapy", "apexpy", "astrometry-azel", "cdflib", "dascutils", "dbprocessing", "dmsp", "enlilviz", "fiasco", "fisspy", "geopack", "georinex", "geospacelab", "ncarglow", "goesutils", "hissw", "hwm93", "igrf", "iri2016", "iri90", "irispy-lmsal", "lowtran", "madrigalWeb", "maidenhead", "mcalf", "mgsutils", "msise00", "ndcube", "nexradutils", "ocbpy", "OMMBV", "pydarn", "pyflct", "pymap3d", "pysatCDF", "pytplot", "pytplot-mpl-temp", "pyzenodo3", "reesaurora", "sciencedates", "SkyWinder", "SkyWinder-Analysis", "solarmach", "solo-epd-loader", "speasy", "spiceypy", "sunkit-image", "sunkit-instruments", "sunraster", "themisasi", "TomograPy", "viresclient", "wmm2015", "wmm2020"]


def is_package_in_pypi(package_name):
    response = requests.get(f'https://pypi.org/pypi/{package_name}/json')
    return response.status_code == 200


def is_package_in_conda(package_name):
    response = requests.get(f'https://api.anaconda.org/package/conda-forge/{package_name}')
    return response.status_code == 200


# ----------------------------------------------------------------------------------------------------------------------

core_packages = get_core_pyhc_packages()
other_packages = get_other_pyhc_packages()
packages = core_packages + other_packages

results = []

for package in packages:
    in_pypi = is_package_in_pypi(package)
    time.sleep(2)  # wait to avoid API rate limit
    in_conda = is_package_in_conda(package)
    time.sleep(2)  # wait to avoid API rate limit

    results.append({
        'Package': package,
        'PyPI': in_pypi,
        'Conda': in_conda,
    })
    print(f"Appended: {package}")

df = pd.DataFrame(results).set_index('Package')  # The pip/conda True/False results

# Use pd options to avoid truncating printout
pd.set_option('display.max_rows', None)
print("\n\nRESULTS:\n")
print(df)
pd.reset_option('display.max_rows')

# first, export the DataFrame to an Excel file
excel_file = "output.xlsx"
df.to_excel(excel_file)

# load the workbook and select the sheet
book = load_workbook(excel_file)

# define the fill colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# get the active worksheet
sheet = book.active

# iterate over the cells in the sheet
for row in sheet.iter_rows(min_row=2):  # start from 2 to skip the header
    for cell in row:
        if cell.value is True:
            cell.fill = green_fill
        elif cell.value is False:
            cell.fill = red_fill

# save the changes
book.save(excel_file)

print("done")
